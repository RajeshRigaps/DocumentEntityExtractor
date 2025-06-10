import os
import io
from flask import Flask, request, jsonify, render_template
import google.generativeai as genai
from docx import Document
import openpyxl
import pandas as pd
from PyPDF2 import PdfReader

# PII Redaction Imports
from presidio_analyzer import AnalyzerEngine
from presidio_anonymizer import AnonymizerEngine
from presidio_anonymizer.entities import OperatorConfig

app = Flask(__name__)

# Configure your Gemini API key
genai.configure(api_key=os.environ.get("GEMINI_API_KEY"))

# Initialize the Gemini model
generation_config = {
    "temperature": 0.2,
    "top_p": 1,
    "top_k": 32,
    "max_output_tokens": 4096,
}

safety_settings = [
    {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
    {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
    {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
    {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
]

model = genai.GenerativeModel(
    model_name="gemini-1.5-flash",
    generation_config=generation_config,
    safety_settings=safety_settings
)

# Initialize Presidio Analyzer and Anonymizer
analyzer = AnalyzerEngine()
anonymizer = AnonymizerEngine()

def anonymize_text_with_presidio(text):
    """
    Detects PII in the given text and anonymizes it using Presidio.
    Replaces detected PII with entity type placeholders (e.g., <PERSON>).
    """
    if not text or not text.strip():
        return ""

    try:
        results = analyzer.analyze(text=text, language='en')

        # Filter out DATE_TIME and IN_PAN entities from the results so they are not anonymized
        filtered_results = [
            result for result in results 
            if result.entity_type != "DATE_TIME" and result.entity_type != "IN_PAN"
        ]

        anonymized_text = anonymizer.anonymize(
            text=text,
            analyzer_results=filtered_results, # Use filtered results
            operators={
                "PERSON": OperatorConfig("replace", {"new_value": "<PERSON>"}),
                "EMAIL_ADDRESS": OperatorConfig("replace", {"new_value": "<EMAIL_ADDRESS>"}),
                "PHONE_NUMBER": OperatorConfig("replace", {"new_value": "<PHONE_NUMBER>"}),
                "STREET_ADDRESS": OperatorConfig("replace", {"new_value": "<ADDRESS>"}),
                "LOCATION": OperatorConfig("replace", {"new_value": "<LOCATION>"}),
                "ORG": OperatorConfig("replace", {"new_value": "<ORGANIZATION>"}),
                # DATE_TIME and IN_PAN are now excluded at the filtering step,
                # so no operator config is needed for them here.
            }
        )
        print("--- Original Text (Snippet) ---")
        print(text[:500] + "...")
        print("\n--- Anonymized Text (PII Redacted Snippet) ---")
        print(anonymized_text.text[:500] + "...")
        return anonymized_text.text
    except Exception as e:
        print(f"Error during PII anonymization: {e}")
        return text

# --- Existing text extraction functions (no changes needed here) ---
def extract_text_from_pdf(file_stream):
    """Extracts text from a PDF file."""
    text = ""
    try:
        reader = PdfReader(file_stream)
        if reader.is_encrypted:
            print("PDF is encrypted. Please provide a non-encrypted PDF.")
            return ""

        if not reader.pages:
            print("PDF has no pages or failed to read pages.")
            return ""

        for page_num, page in enumerate(reader.pages):
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
            else:
                print(f"Warning: No text extracted from page {page_num + 1}. This might be an image-based page.")
        
        if not text.strip():
            print("No text extracted from the entire PDF. It might be entirely image-based or empty.")

    except PdfReadError as e:
        print(f"PyPDF2 Read Error: {e}. The PDF might be corrupted or malformed.")
    except Exception as e:
        print(f"General Error extracting text from PDF: {e}")
    return text

def extract_text_from_docx(file_stream):
    """Extracts text from a DOCX file."""
    text = ""
    try:
        document = Document(file_stream)
        for paragraph in document.paragraphs:
            text += paragraph.text + "\n"
    except Exception as e:
        print(f"Error extracting text from DOCX: {e}")
    return text

def extract_text_from_xlsx(file_stream):
    """Extracts text from an XLSX file."""
    text = ""
    try:
        workbook = openpyxl.load_workbook(file_stream)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        text += str(cell.value) + " "
                text += "\n"
    except Exception as e:
        print(f"Error extracting text from XLSX: {e}")
    return text

def extract_text_from_csv(file_stream):
    """Extracts text from a CSV file."""
    try:
        df = pd.read_csv(file_stream)
        return df.to_string(index=False)
    except Exception as e:
        print(f"Error extracting text from CSV: {e}")
        return ""

@app.route('/')
def index():
    """Renders the upload form."""
    return render_template('index.html')

@app.route('/extract', methods=['POST'])
def extract_entities():
    """Handles file upload and entity extraction."""
    if 'document' not in request.files:
        return jsonify({"error": "No document part in the request"}), 400

    file = request.files['document']

    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    if file:
        file_extension = file.filename.split('.')[-1].lower()
        file_stream = io.BytesIO(file.read())
        extracted_text = ""

        if file_extension == 'pdf':
            extracted_text = extract_text_from_pdf(file_stream)
        elif file_extension == 'docx':
            extracted_text = extract_text_from_docx(file_stream)
        elif file_extension == 'xlsx':
            extracted_text = extract_text_from_xlsx(file_stream)
        elif file_extension == 'csv':
            extracted_text = extract_text_from_csv(file_stream)
        else:
            return jsonify({"error": "Unsupported file type"}), 400

        if not extracted_text.strip():
            return jsonify({"error": "Could not extract text from the document. The document might be empty or in an unsupported format."}), 400

        redacted_text = anonymize_text_with_presidio(extracted_text)

        # --- REFINED PROMPT FOR DIRECT KEY-VALUE PAIRS ---
        prompt = f"""
        You are an expert in extracting structured information from documents.
        Extract all relevant entities from the following document text and return them as a single JSON object.
        Each entity should be a direct key-value pair within the JSON object.

        - For a clear "Key: Value" pair (e.g., "Job Title: Software Engineer"), use the "Key" as the JSON field name and the "Value" as its corresponding value.
        - For a standalone entity (e.g., a company name), choose an appropriate field name (e.g., "Company Name") and use the entity as its value.
        - IMPORTANT for PII: If an entity in the document text is a redacted PII placeholder (e.g., <PERSON>, <EMAIL_ADDRESS>, <PHONE_NUMBER>, <ADDRESS>, <LOCATION>, <ORGANIZATION>), you MUST extract it.
            - If it comes with an explicit label (e.g., "Employee Name: <PERSON>"), use that label as the JSON field name (e.g., "Employee Name") and the placeholder as its value (e.g., "<PERSON>").
            - If it's a standalone placeholder (e.g., just "<EMAIL_ADDRESS>"), choose an appropriate field name (e.g., "Contact Email") and use the placeholder itself as its value (e.g., "<EMAIL_ADDRESS>").
            - Do NOT try to infer or recreate the original PII. The placeholder IS the value.
        - IMPORTANT for Dates: Date entities will *not* be redacted. Extract them directly as they appear.

        If there are multiple distinct records (e.g., multiple employees in a timesheet), generate a JSON array of objects, where each object represents a record. If it's a single document with overall entities (like an offer letter), generate a single JSON object.

        Examples of desired JSON structure:
        // For a single document (e.g., Offer Letter):
        {{
            "Job Title": "Software Engineer",
            "Company Name": "<ORGANIZATION>",
            "Employee Name": "<PERSON>",
            "Start Date": "July 7, 2025", // DATE is now extracted as plain text
            "Contact Email": "<EMAIL_ADDRESS>",
            "Company Address": "<ADDRESS>",
            "Office Location": "<LOCATION>",
            "Company Phone": "<PHONE_NUMBER>",
            "Base Salary": "$125,000 per year"
        }}

        // For multiple records (e.g., a Timesheet Summary with multiple entries):
        [
            {{
                "Employee Name": "<PERSON>",
                "Total Hours": "40",
                "Week Ending": "June 7, 2025" // DATE is now extracted as plain text
            }},
            {{
                "Employee Name": "<PERSON>",
                "Total Hours": "38",
                "Week Ending": "June 14, 2025" // DATE is now extracted as plain text
            }}
        ]

        Ensure the output is a valid JSON object or array of objects.

        Document Text:
        ---
        {redacted_text}
        ---

        JSON Output:
        """
        # --- END REFINED PROMPT FOR DIRECT KEY-VALUE PAIRS ---

        try:
            response = model.generate_content(prompt)
            
            response_text = response.text.strip()

            if response_text.startswith("```json") and response_text.endswith("```"):
                json_string = response_text[7:-3].strip()
            else:
                json_string = response_text

            import json
            entities = json.loads(json_string)
            return jsonify(entities)
        except Exception as e:
            print(f"Error calling Gemini API or parsing JSON: {e}")
            return jsonify({"error": "Failed to extract entities. Please check the document content or try again later.", "details": str(e)}), 500

    return jsonify({"error": "An unknown error occurred"}), 500

if __name__ == '__main__':
    if not os.environ.get("GEMINI_API_KEY"):
        print("WARNING: GEMINI_API_KEY environment variable not set.")
        print("Please set it before running the application.")
        print("Example: export GEMINI_API_KEY='your_api_key_here'")
    app.run(debug=True)